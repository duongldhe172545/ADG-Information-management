import openpyxl

def main():
    file_path = r'd:\ADG-Dealer\KhachHang_TatCa_2026-03-30 (1).xlsx'
    out_path = r'd:\ADG-Dealer\KhachHang_TatCa_2026-03-30_Cleaned_With_Styles.xlsx'
    
    print("Đang đọc file Excel (giữ nguyên định dạng và dropdown)...")
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Find SĐT column
    header_row = 1
    sdt_col = None
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col).value
        if val == 'SĐT':
            sdt_col = col
            break

    if not sdt_col:
        print("Lỗi: Không tìm thấy cột SĐT!")
        return

    seen_phones = set()
    rows_to_delete = []

    print(f"Đang phân tích {sheet.max_row} dòng...")
    for row_idx in range(2, sheet.max_row + 1):
        # 1. Kiểm tra dòng trống hoàn toàn
        is_empty = True
        for col in range(1, sheet.max_column + 1):
            c_val = sheet.cell(row=row_idx, column=col).value
            if c_val is not None and str(c_val).strip() != '':
                is_empty = False
                break
                
        if is_empty:
            rows_to_delete.append(row_idx)
            continue
            
        # 2. Lấy SĐT
        sdt = sheet.cell(row=row_idx, column=sdt_col).value
        sdt_str = str(sdt).strip() if sdt is not None else ''
        
        # 3. Nếu không có SĐT
        if not sdt_str or sdt_str.lower() == 'nan':
            rows_to_delete.append(row_idx)
            continue
            
        # 4. Kiểm tra trùng
        if sdt_str in seen_phones:
            rows_to_delete.append(row_idx)
        else:
            seen_phones.add(sdt_str)

    print(f"Phát hiện được {len(rows_to_delete)} dòng lỗi/trùng. Cần phải xóa.")
    print("Đang xóa dòng từ dưới lên (để giữ layout và dropdown)...")
    
    # Delete from bottom to top to preserve index order
    for idx, r in enumerate(reversed(rows_to_delete)):
        sheet.delete_rows(r)
        if idx % 100 == 0 and idx > 0:
            print(f" Đã xóa {idx} dòng...")

    print("Đang lưu lại file...")
    wb.save(out_path)
    print("====================")
    print(f"Hoàn tất! Đã lưu file sạch tại: {out_path}")

if __name__ == '__main__':
    main()
