"""Export service — multi-sheet Excel export with formatting."""

import os
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database.db import get_db
from config import EXPORT_FOLDER

VN_TZ = timezone(timedelta(hours=7))

# ─── Column Definitions ─────────────────────────────────────────
DATA_COLUMNS = [
    ('stt', 'STT', 10),
    ('loai_kh', 'Loại KH', 15),
    ('ten_cong_ty', 'Tên Công Ty/Đơn Vị', 30),
    ('ho_ten', 'Họ Tên Chủ', 22),
    ('sdt', 'SĐT', 18),
    ('tinh', 'Tỉnh', 15),
    ('xa', 'Xã', 15),
    ('dia_chi', 'Địa Chỉ', 30),
    ('nguon', 'Nguồn', 15),
    ('khu_vuc', 'Khu vực', 12),
    ('so_kh_quay_lai', 'Số KH quay lại/năm', 12),
    ('biet_loi_nhuan', 'Biết lợi nhuận từng đơn?', 20),
    ('doi_tho', 'Đội thợ thi công', 18),
    ('chinh_sach_bh', 'Chính sách BH với khách', 20),
    ('muc_quan_tam', 'Mức quan tâm hợp tác', 20),
    ('ban_kinh_km', 'Bán kính KH gọi đến (km)', 12),
    ('quan_ly_data', 'Cách quản lý data KH', 20),
    ('kiem_soat_mua_hang', 'Kiểm soát mua hàng', 20),
    ('so_nguoi_gioi_thieu', 'Số người được giới thiệu', 12),
    ('thong_tin_chi_tiet', 'Thông Tin Chi Tiết', 35),
]

SCORE_COLUMNS = [
    ('c1', 'Điểm C1 (Quay lại)', 10),
    ('c2', 'Điểm C2 (Lợi nhuận)', 10),
    ('c3', 'Điểm C3 (Đội thợ)', 10),
    ('c4', 'Điểm C4 (Chính sách BH)', 10),
    ('c5', 'Điểm C5 (Quan tâm)', 10),
    ('c6', 'Điểm C6 (Bán kính)', 10),
    ('c7', 'Điểm C7 (Quản lý data)', 10),
    ('c8', 'Điểm C8 (Mua hàng)', 10),
    ('c9', 'Điểm C9 (Giới thiệu)', 10),
    ('c_score', 'TỔNG ĐIỂM', 10),
    ('tier', 'HẠNG', 8),
]

ALL_EXPORT_COLS = DATA_COLUMNS + SCORE_COLUMNS

LOG_COLUMNS = [
    ('stt', 'STT', 6),
    ('timestamp', 'Thời gian', 20),
    ('action_type', 'Loại hành động', 18),
    ('detail', 'Chi tiết', 40),
    ('customer_id', 'Mã KH', 8),
    ('customer_name', 'Tên KH', 25),
    ('operator', 'Người thao tác', 15),
]

# ─── Styles ──────────────────────────────────────────────────────
FONT_BOLD = Font(name='Arial', size=10, bold=True)
FONT_NORMAL = Font(name='Arial', size=10)
FONT_TITLE = Font(name='Arial', size=12, bold=True)

FILL_BLUE = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
FILL_ORANGE = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
FILL_GREEN = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'),
    top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6')
)


def _write_header(ws, columns, fills_map):
    """Write header row with styles."""
    for col_idx, (db_key, header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = FONT_BOLD
        cell.fill = fills_map.get(db_key, FILL_BLUE)
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


def _build_sheet_customers(wb, rows):
    """Sheet 1: Dữ Liệu Khách Hàng."""
    ws = wb.active
    ws.title = 'Dữ Liệu Khách Hàng'

    # Header fill map
    fills = {}
    for db_key, _, _ in DATA_COLUMNS:
        fills[db_key] = FILL_BLUE
    for db_key, _, _ in SCORE_COLUMNS:
        fills[db_key] = FILL_ORANGE

    _write_header(ws, ALL_EXPORT_COLS, fills)

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        data = dict(row)
        for col_idx, (db_key, _, _) in enumerate(ALL_EXPORT_COLS, 1):
            val = ''
            if db_key == 'stt':
                val = row_idx - 1
            else:
                val = data.get(db_key, '')

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER if db_key in (
                'stt', 'c1', 'c2', 'c3', 'c4', 'c5', 'c6', 'c7', 'c8', 'c9', 'c_score', 'tier'
            ) else ALIGN_LEFT

            # Tier highlights
            if db_key == 'tier' and val:
                if val in ('A', 'B'):
                    cell.fill = FILL_GREEN
                    cell.font = Font(name='Arial', size=10, bold=True, color='385623')
                elif val in ('C', 'D'):
                    cell.fill = FILL_GRAY
                    cell.font = Font(name='Arial', size=10, bold=True, color='7F7F7F')

    ws.freeze_panes = 'F2'
    return ws


def _build_sheet_log(wb):
    """Sheet 2: Nhật Ký Thay Đổi."""
    ws = wb.create_sheet('Nhật Ký Thay Đổi')
    
    fills = {k: FILL_GREEN for k, _, _ in LOG_COLUMNS}
    _write_header(ws, LOG_COLUMNS, fills)

    conn = get_db()
    logs = conn.execute("SELECT * FROM change_log ORDER BY id DESC").fetchall()
    conn.close()

    for row_idx, log in enumerate(logs, 2):
        data = dict(log)
        for col_idx, (db_key, _, _) in enumerate(LOG_COLUMNS, 1):
            if db_key == 'stt':
                val = row_idx - 1
            elif db_key == 'timestamp':
                val = _utc_to_vn(data.get('timestamp', ''))
            else:
                val = data.get(db_key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_LEFT

    ws.freeze_panes = 'A2'
    return ws


def _build_sheet_dashboard(wb):
    """Sheet 3: Tổng Hợp (Dashboard)."""
    ws = wb.create_sheet('Tổng Hợp')
    conn = get_db()

    total = conn.execute("SELECT COUNT(*) FROM customers WHERE is_deleted=0").fetchone()[0]
    scored = conn.execute("SELECT COUNT(*) FROM customers WHERE is_deleted=0 AND c_score > 0").fetchone()[0]
    needs_review = conn.execute("SELECT COUNT(*) FROM customers WHERE is_deleted=0 AND review_status='needs_review'").fetchone()[0]
    deleted = conn.execute("SELECT COUNT(*) FROM customers WHERE is_deleted=1").fetchone()[0]

    # ─── Bảng 1: Tổng quan ───
    ws.cell(row=1, column=1, value='TỔNG QUAN').font = FONT_TITLE
    summary = [
        ('Tổng KH', total),
        ('Đã chấm điểm', scored),
        ('Cần review', needs_review),
        ('KH đã xoá', deleted),
    ]
    for i, (label, val) in enumerate(summary, 2):
        a = ws.cell(row=i, column=1, value=label)
        a.font = FONT_BOLD; a.fill = FILL_YELLOW; a.border = THIN_BORDER
        b = ws.cell(row=i, column=2, value=val)
        b.font = FONT_NORMAL; b.border = THIN_BORDER; b.alignment = ALIGN_CENTER

    # ─── Bảng 2: Phân bổ Tier ───
    ws.cell(row=7, column=1, value='PHÂN BỔ TIER').font = FONT_TITLE
    for col_idx, h in enumerate(['Tier', 'Số lượng', 'Tỷ lệ %'], 1):
        c = ws.cell(row=8, column=col_idx, value=h)
        c.font = FONT_BOLD; c.fill = FILL_YELLOW; c.border = THIN_BORDER; c.alignment = ALIGN_CENTER

    tiers = conn.execute(
        "SELECT tier, COUNT(*) FROM customers WHERE is_deleted=0 AND tier != '' GROUP BY tier ORDER BY tier"
    ).fetchall()

    for i, (tier, count) in enumerate(tiers, 9):
        pct = round(count / total * 100, 1) if total else 0
        for col_idx, val in enumerate([tier, count, pct], 1):
            c = ws.cell(row=i, column=col_idx, value=val)
            c.font = FONT_NORMAL; c.border = THIN_BORDER; c.alignment = ALIGN_CENTER

    # ─── Bảng 3: Top 10 Tỉnh ───
    r = 9 + len(tiers) + 1
    ws.cell(row=r, column=1, value='TOP 10 TỈNH').font = FONT_TITLE
    r += 1
    for col_idx, h in enumerate(['Tỉnh', 'Số KH'], 1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = FONT_BOLD; c.fill = FILL_YELLOW; c.border = THIN_BORDER; c.alignment = ALIGN_CENTER
    r += 1

    top_tinh = conn.execute(
        "SELECT tinh, COUNT(*) AS cnt FROM customers WHERE is_deleted=0 AND tinh IS NOT NULL AND tinh != '' GROUP BY tinh ORDER BY cnt DESC LIMIT 10"
    ).fetchall()
    for tinh, cnt in top_tinh:
        ws.cell(row=r, column=1, value=tinh).font = FONT_NORMAL
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=cnt).font = FONT_NORMAL
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER
        r += 1

    # ─── Bảng 4: Phân bổ Nguồn ───
    r += 1
    ws.cell(row=r, column=1, value='PHÂN BỔ NGUỒN').font = FONT_TITLE
    r += 1
    for col_idx, h in enumerate(['Nguồn', 'Số KH'], 1):
        c = ws.cell(row=r, column=col_idx, value=h)
        c.font = FONT_BOLD; c.fill = FILL_YELLOW; c.border = THIN_BORDER; c.alignment = ALIGN_CENTER
    r += 1

    sources = conn.execute(
        "SELECT nguon, COUNT(*) AS cnt FROM customers WHERE is_deleted=0 AND nguon IS NOT NULL AND nguon != '' GROUP BY nguon ORDER BY cnt DESC"
    ).fetchall()
    for nguon, cnt in sources:
        ws.cell(row=r, column=1, value=nguon).font = FONT_NORMAL
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=cnt).font = FONT_NORMAL
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER
        r += 1

    conn.close()

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    return ws


def _utc_to_vn(ts_str):
    """Convert UTC timestamp string to VN timezone string."""
    if not ts_str:
        return ''
    try:
        dt = datetime.strptime(str(ts_str), '%Y-%m-%d %H:%M:%S')
        dt = dt.replace(tzinfo=timezone.utc).astimezone(VN_TZ)
        return dt.strftime('%d/%m/%Y %H:%M:%S')
    except (ValueError, TypeError):
        return str(ts_str)


def export_customers(filters=None, filename=None):
    """Export customers to multi-sheet Excel. Returns filepath."""
    if not filename:
        now = datetime.now(VN_TZ)
        filename = f"KhachHang_ADG_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    conn = get_db()
    query = "SELECT * FROM customers WHERE is_deleted = 0"
    params = []

    if filters:
        filter_keys = [c[0] for c in ALL_EXPORT_COLS]
        for key, value in filters.items():
            if value and key in filter_keys:
                query += f" AND {key} = ?"
                params.append(value)

    query += " ORDER BY id ASC"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()

    # Sheet 1: Data
    _build_sheet_customers(wb, rows)

    # Sheet 2: Log
    _build_sheet_log(wb)

    # Sheet 3: Dashboard
    _build_sheet_dashboard(wb)

    filepath = os.path.join(EXPORT_FOLDER, filename)
    wb.save(filepath)
    return filepath
