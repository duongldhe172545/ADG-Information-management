import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

target_file = r"d:\ADG-Dealer\KhachHang_Template_Chuan.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DS_KhachHang"

# 19 Columns matching HEADER_MAP exactly
headers = [
    'Loại KH', 
    'Tên Công Ty/Đơn Vị', 
    'Họ Tên Chủ', 
    'SĐT', 
    'Tỉnh', 
    'Xã', 
    'Địa Chỉ', 
    'Nguồn', 
    'Thông Tin Chi Tiết', 
    'Khu vực', 
    'Số KH quay lại/năm', 
    'Biết lợi nhuận từng đơn?', 
    'Đội thợ thi công', 
    'Chính sách BH với khách', 
    'Mức quan tâm hợp tác', 
    'Bán kính KH gọi đến (km)', 
    'Cách quản lý data KH', 
    'Kiểm soát mua hàng', 
    'Số người được giới thiệu'
]

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
align_center = Alignment(horizontal='center', vertical='center')

for col_idx, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx, value=h)
    c.fill = header_fill
    c.font = header_font
    c.border = thin_border
    c.alignment = align_center

# Widths based on new columns
ws.column_dimensions['A'].width = 20 # Loai KH
ws.column_dimensions['B'].width = 35 # Ten...
ws.column_dimensions['C'].width = 20 # Ho ten...
ws.column_dimensions['D'].width = 15 # SDT
ws.column_dimensions['E'].width = 15 # Tinh
ws.column_dimensions['F'].width = 15 # Xa
ws.column_dimensions['G'].width = 40 # Dia chi
ws.column_dimensions['H'].width = 15 # Nguon
ws.column_dimensions['I'].width = 40 # Thong tin chi tiet
ws.column_dimensions['J'].width = 15 # Khu vuc

for col_char in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
    ws.column_dimensions[col_char].width = 30

# Validations
dv_c2 = DataValidation(type="list", formula1='"Không biết,Biết LN nhưng DSO>60 ngày,Biết LN>15% và DSO<=60 ngày,Khác"', allow_blank=True)
dv_c3 = DataValidation(type="list", formula1='"Không có đội,1-3 thợ rời theo vụ,>=2 thợ cơ hữu SLA ổn,Khác"', allow_blank=True)
dv_c4 = DataValidation(type="list", formula1='"Đổ lỗi NCC,BH nhưng đòi hoàn NCC,Tự ký BH chịu CP,Khác"', allow_blank=True)
dv_c5 = DataValidation(type="list", formula1='"Không muốn đổi,Quan tâm chưa rõ lợi ích,Có nỗi đau cụ thể muốn giải,Khác"', allow_blank=True)
dv_c7 = DataValidation(type="list", formula1='"Không ghi chép,Ghi Zalo/Excel rải rác,Có hệ thống xuất được lịch sử,Khác"', allow_blank=True)
dv_c8 = DataValidation(type="list", formula1='"Theo chỉ định NCC,Có 2-3 NCC lựa chọn,Chủ động thương lượng giá,Khác"', allow_blank=True)

ws.add_data_validation(dv_c2)
ws.add_data_validation(dv_c3)
ws.add_data_validation(dv_c4)
ws.add_data_validation(dv_c5)
ws.add_data_validation(dv_c7)
ws.add_data_validation(dv_c8)

# Map to the exactly index:
# K=11(C1), L=12(C2), M=13(C3), N=14(C4), O=15(C5), P=16(C6), Q=17(C7), R=18(C8), S=19(C9)
dv_c2.add("L2:L1000")
dv_c3.add("M2:M1000")
dv_c4.add("N2:N1000")
dv_c5.add("O2:O1000")
dv_c7.add("Q2:Q1000")
dv_c8.add("R2:R1000")

# Delete old template to avoid locked file errors
try:
    import os
    if os.path.exists(target_file):
        os.remove(target_file)
except:
    pass

try:
    wb.save(target_file)
    print(f"File saved successfully to {target_file}")
except Exception as e:
    # Fallback name if permission denied
    target_file = r"d:\ADG-Dealer\KhachHang_Template_Chuan_v2.xlsx"
    wb.save(target_file)
    print(f"File saved to new name '{target_file}' due to old file being locked.")

