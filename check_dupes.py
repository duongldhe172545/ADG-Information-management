import openpyxl
import re
from collections import defaultdict

def extract_all_phones(phone_str):
    if not phone_str:
        return []
    s = str(phone_str).replace(' ', '').replace('.', '').replace('-', '')
    parts = re.split(r'[/,;]', s)
    res = []
    for p in parts:
        p = re.sub(r'\D', '', p)
        if p:
            if p.startswith('84'):
                p = '0' + p[2:]
            elif p.startswith('+84'):
                p = '0' + p[3:]
            res.append(p)
    return res

def main():
    print("Đang đọc file Excel để kiểm tra trùng chéo (bỏ qua dấu cách, chấm, đuôi phụ)...")
    wb = openpyxl.load_workbook(r'd:\ADG-Dealer\KhachHang_TatCa_2026-03-30_Cleaned_With_Styles.xlsx', read_only=True)
    sheet = wb.active

    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(r).strip() if r else '' for r in row]
    
    try:
        sdt_idx = headers.index('SĐT')
        name_idx = headers.index('Tên Công Ty/Đơn Vị') if 'Tên Công Ty/Đơn Vị' in headers else -1
    except ValueError:
        print("Lỗi: Không tìm thấy cột SĐT!")
        return

    all_phones = defaultdict(list)
    row_count = 2

    for row in sheet.iter_rows(min_row=2, values_only=True):
        val = row[sdt_idx]
        if val:
            phones = extract_all_phones(val)
            for p in phones:
                all_phones[p].append(row_count)
        row_count += 1

    dupes = {p: rows for p, rows in all_phones.items() if len(rows) > 1}

    print(f"\nPhát hiện: {len(dupes)} số điện thoại vẫn bị lặp lại (nằm ở {sum(len(r) for r in dupes.values())} dòng) do format gõ chữ/phím khác nhau.")
    
    if dupes:
        print("\n=== VÍ DỤ 5 SĐT BỊ TRÙNG DO FORMAT KHÁC  ===")
        count = 0
        for p, rows in list(dupes.items()):
            print(f"- Số chuẩn: {p} => Xuất hiện chéo ở các dòng: {rows}")
            count += 1
            if count >= 5:
                break
if __name__ == '__main__':
    main()
